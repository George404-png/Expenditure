from tkinter.ttk import Style

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Color, Border, Side, colors, PatternFill, Font

#Input the location of the csv file in CSV file and the location where the excel sheet should be created
# an example of this is '/Users/name/Desktop/statement.csv
CSV_file = ''
Outputfile = ''

data = pd.read_csv('CSV_file', decimal=',',sep=';')
print(data.head())

newdata = data.drop(['Account', 'Counterparty', 'Code', 'Transaction type', 'Notifications','Resulting balance','Tag'], axis=1)
print(newdata.head())
newdata['Amount (EUR)'] = pd.to_numeric(newdata['Amount (EUR)'])
newdata.replace('Debit',1)
newdata.replace('Credit', -1)
print(newdata.dtypes)
Category_Collectrion = {}
food = pd.DataFrame(columns=newdata.columns)
movies = pd.DataFrame(columns=newdata.columns)
transport = pd.DataFrame(columns=newdata.columns)
spotify = pd.DataFrame(columns=newdata.columns)
mobile = pd.DataFrame(columns=newdata.columns)
other = pd.DataFrame(columns=newdata.columns)
count = [0, 0, 0, 0, 0, 0]

#Within the Keywords food array input a list of names that the program will look for in order to distinguish this entry as a food item
#An example of this is ['Albert Heijn', 'Plus']

keywords_food = []
Categories = ["Food", "Movies", "Transport", "Spotify", "Mobile", "Other"]
ExcemptKey = []
yellow_back = PatternFill(bgColor=colors.YELLOW)
numberofrows = 6
tablecounter = 2


def classifier1(x):
    global count
    if any(word in x[1] for word in keywords_food):
        food.loc[count[0]] = x
        count[0] += 1
    elif "Pathe" in x[1]:
        movies.loc[count[1]] = x
        count[1] += 1
    elif "NS GROEP" in x[1]:
        transport.loc[count[2]] = x
        count[2] += 1
    elif "SPOTIFY" in x[1]:
        spotify.loc[count[3]] = x
        count[3] += 1
    elif "TELE2" in x[1]:
        mobile.loc[count[4]] = x
        count[4] += 1
    elif any(word in x[1] for word in ExcemptKey):
        nothing = 0
    else:
        other.loc[count[5]] = x
        count[5] += 1


def TableDrawer(sheet, week, ):
    wordcounter = 0
    for j in range(1, 9):
        for i in range(1, 6):
            if j == 1 and i == 1:
                sheet.cell(i, j, "WEEKS").font = Font(bold=True)
            elif j == 8 and i == 1:
                sheet.cell(i, j, "Total Monthly").font = Font(bold=True)
            elif i == 1 and j != 1 and j != 8:
                sheet.cell(i, j, Categories[wordcounter]).font = Font(bold=True)
                wordcounter += 1
            elif i == 3 and j == 1:
                sheet.cell(i, j, "WK1-4").font = Font(bold=True)
            elif i == 4 and j == 1:
                sheet.cell(i, j, "WK3-4").font = Font(bold=True)
            elif i == 5 and j == 1:
                sheet.cell(i, j, "Total").font = Font(bold=True)

            if i == 3:
                sheet.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
            elif i == 4:
                sheet.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=colors.BLUE)
            elif i == 1 or i == 2:
                sheet.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=colors.GREEN)

            sheet.cell(i, j).border = Border(left=Side(border_style="medium"), right=Side(border_style="medium"),
                                             top=Side(border_style="medium"), bottom=Side(border_style="medium"))
        sheet.merge_cells(start_column=j, start_row=1, end_column=j, end_row=2)


def excelwriter():
    workbook = Workbook()
    sheet = workbook.active

    TableDrawer(sheet, 1)

    for i in range(1, 6):
        if i == 1:
            sheet.cell(7, i, "WEEK 1-4").font = Font(bold=True)

        sheet.cell(7, 1).fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)

    sheet.merge_cells(start_column=1, start_row=7, end_column=5, end_row=7)

    exportdata(numberofrows + 2, 2, food, sheet, Categories[0])
    exportdata(numberofrows + 2, 2, movies, sheet, Categories[1])
    exportdata(numberofrows + 2, 2, transport, sheet, Categories[2])
    exportdata(numberofrows + 2, 2, spotify, sheet, Categories[3])
    exportdata(numberofrows + 2, 2, mobile, sheet, Categories[4])
    exportdata(numberofrows + 2, 2, other, sheet, Categories[5])

    sheet['H3'] = '=SUM(B3:G3)'

    workbook.save(filename=Outputfile)


def exportdata(startrow, startcol, x, sheet, name):
    counter2 = 0
    global tablecounter
    for row1 in x.iloc:
        for i in range(0, 4):
            number = row1[i]
            if row1[2] == 'Credit' and i == 3:
                number = -1*row1[3]
            else:
                number = row1[i]
            sheet.cell(startrow + counter2, startcol + i, number)
        counter2 += 1
    for i in range(1, 7):
        sheet.cell(startrow + counter2, i).border = Border(top=Side(border_style="medium"))
    sheet.cell(startrow + counter2, 1, "Week 1-2").fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
    sheet.cell(startrow + counter2, 5, name).fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)

    global numberofrows
    numberofrows = startrow + counter2
    sheet.cell(startrow + counter2, 6,
               '=SUM(E' + str(startrow) + ':E' + str(numberofrows - 1) + ')').fill = PatternFill(fill_type='solid',
                                                                                                 fgColor=colors.YELLOW)
    sheet.cell(3, tablecounter, '=SUM(E' + str(startrow) + ':E' + str(numberofrows - 1) + ')')
    tablecounter += 1


for row in newdata.iloc:
    classifier1(row)
excelwriter()
